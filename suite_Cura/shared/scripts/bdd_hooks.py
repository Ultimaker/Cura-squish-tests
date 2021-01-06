# -*- coding: utf-8 -*-
import os
import json
import squish
import datetime
import pandas as pd
from openpyxl import load_workbook


@OnFeatureEnd
def hook(context):
    for ctx in applicationContextList():
        ctx.detach()


@OnScenarioStart
def hook(context):
    if context.userData:
        context.userData.clear()


# TODO: Fix when updating Performance tests to export results
@OnScenarioEnd
def hook(context, current_data=None):
    if current_data is None:
        current_data = {}

    if context.userData:
        # Here we declare the file we want to use to store the performance output of our tests
        performance_results = findFile("scripts", "TestResults/Performance.xlsx")

        # Get only the extension of the file
        f_base, f_ext = os.path.splitext(performance_results)
        
        cura_aut = currentApplicationContext()
        aut_path = cura_aut.cwd
        
        if f_ext == ".xlsx":
            # read the excel sheet in case it has values already inserted
            try:
                df = pd.read_excel(performance_results, engine="openpyxl")
                now = datetime.datetime.now()
                exceldata = pd.DataFrame({'Measurement': [next(iter(context.userData.values()))], 'Datetime': [now.strftime("%Y-%m-%d %H:%M:%S")], 'Cura Build': [os.path.basename(os.path.normpath(aut_path))]}, index=[context.title])

                if "Measurement" in df:
                    append_df_to_excel(performance_results, exceldata)
                else:
                    with pd.ExcelWriter(performance_results) as writer:
                        exceldata.to_excel(writer)
                        
            except Exception as e:
                print(e)
            
        else:
            with open(performance_results, "r") as new_file:
                try:
                    current_data = json.load(new_file)
                except Exception as e:
                    print(e)

            # Format data so its more readable and easier to analyze
            jsondata = {"data": []}
            content, results = {}, {}

            # Using context.userData in steps to store results that I can extract here
            # TODO: context.userData is sometimes used for things other than performance results. Filter those out!
            results["GUI"] = next(iter(context.userData.values()))

            content["measurements"] = context.title
            now = datetime.datetime.now()
            content["datetime"] = now.strftime("%Y-%m-%d %H:%M:%S")
            content["fields"] = results

            with open(performance_results, "w") as new_file:
                if "data" in current_data:
                    current_data['data'].append(content)
                    json.dump(current_data, new_file, indent=4)
                else:
                    jsondata['data'].append(content)
                    json.dump(jsondata, new_file, indent=4)

def init():
    testSettings.logScreenshotOnError = True
    testSettings.logScreenshotOnFail = True


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs, header=False)

    # save the workbook
    writer.save()
